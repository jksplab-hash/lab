import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import io
import os
import json
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable

st.set_page_config(page_title="Screen Printing R&D Portal", layout="wide", page_icon="🎨")

EXCEL_FILE = "Screen_Printing_RND_Technical_Library_Workbook.xlsx"
INK_LIBRARY_SHEET = "Chemical Ink Library"
RECIPE_LOG_SHEET = "Saved Technical Recipes"

# Initial Default Data with Unit Prices ($/kg)
INITIAL_INK_LIBRARY = [
    {"Product Name": "Silicone Base Clear", "Supplier Code": "SIL-BASE-90", "Role": "Base Transparent", "Unit Price ($/kg)": 14.50, "Notes": "High elasticity silicone base"},
    {"Product Name": "Silicone White Undercoat", "Supplier Code": "SIL-WHT-10", "Role": "Base Opaque / White", "Unit Price ($/kg)": 12.00, "Notes": "Provides opacity & bleed barrier"},
    {"Product Name": "High Fastness Black Pigment", "Supplier Code": "PIG-BLK-05", "Role": "Pigment Colorant", "Unit Price ($/kg)": 18.00, "Notes": "Color shade matching"},
    {"Product Name": "High Fastness Red Pigment", "Supplier Code": "PIG-RED-02", "Role": "Pigment Colorant", "Unit Price ($/kg)": 22.00, "Notes": "Vibrant red shade matching"},
    {"Product Name": "Silicone Platinum Catalyst", "Supplier Code": "CAT-SIL-02", "Role": "Catalyst / Hardener", "Unit Price ($/kg)": 45.00, "Notes": "Mix thoroughly before printing"},
    {"Product Name": "Anti-Fading Crosslinker", "Supplier Code": "XL-MOD-01", "Role": "Crosslinker / Fixer", "Unit Price ($/kg)": 28.00, "Notes": "Enhances wash fastness (20+ cycles)"},
    {"Product Name": "Water-Based Elastic Clear Base", "Supplier Code": "WB-BASE-01", "Role": "Base Transparent", "Unit Price ($/kg)": 6.50, "Notes": "Eco-friendly soft hand base"},
    {"Product Name": "Plastisol High-Opacity White", "Supplier Code": "PL-WHT-99", "Role": "Base Opaque / White", "Unit Price ($/kg)": 8.00, "Notes": "Heavy coverage underbase"}
]

# Helper Function: Load Persistent Ink Library from Excel
def load_ink_library():
    if os.path.exists(EXCEL_FILE):
        try:
            xls = pd.ExcelFile(EXCEL_FILE)
            if INK_LIBRARY_SHEET in xls.sheet_names:
                df = pd.read_excel(xls, INK_LIBRARY_SHEET)
                if not df.empty:
                    if "Unit Price ($/kg)" not in df.columns:
                        df["Unit Price ($/kg)"] = 10.0
                    return df
        except Exception as e:
            st.error(f"Error loading chemical library: {e}")
    
    df_default = pd.DataFrame(INITIAL_INK_LIBRARY)
    save_ink_library_to_excel(df_default)
    return df_default

# Helper Function: Save Permanent Ink Library to Excel
def save_ink_library_to_excel(df):
    try:
        if os.path.exists(EXCEL_FILE):
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=INK_LIBRARY_SHEET, index=False)
        else:
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=INK_LIBRARY_SHEET, index=False)
        return True
    except Exception as e:
        st.error(f"Failed to save library to Excel: {e}")
        return False

# Helper Function: Load All Saved Recipes
def load_saved_recipes():
    if os.path.exists(EXCEL_FILE):
        try:
            xls = pd.ExcelFile(EXCEL_FILE)
            if RECIPE_LOG_SHEET in xls.sheet_names:
                return pd.read_excel(xls, RECIPE_LOG_SHEET)
        except Exception as e:
            st.error(f"Error loading recipes: {e}")
    return pd.DataFrame()

# Helper Function: Save New Recipe to Excel
def save_recipe_to_excel(recipe_data):
    try:
        data_to_save = recipe_data.copy()
        data_to_save['formulation'] = json.dumps(recipe_data['formulation'])

        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = RECIPE_LOG_SHEET
            ws.append(list(data_to_save.keys()))
            ws.append(list(data_to_save.values()))
            wb.save(EXCEL_FILE)
        else:
            df_existing = load_saved_recipes()
            new_row = pd.DataFrame([data_to_save])
            df_updated = pd.concat([df_existing, new_row], ignore_index=True) if not df_existing.empty else new_row
            
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_updated.to_excel(writer, sheet_name=RECIPE_LOG_SHEET, index=False)
        return True
    except Exception as e:
        st.error(f"Failed to save recipe to Excel: {e}")
        return False

# Helper Function: Delete a Specific Recipe from Excel
def delete_recipe_from_excel(recipe_id_to_delete):
    try:
        df_recipes = load_saved_recipes()
        if not df_recipes.empty:
            df_updated = df_recipes[df_recipes['recipe_id'] != recipe_id_to_delete].reset_index(drop=True)
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_updated.to_excel(writer, sheet_name=RECIPE_LOG_SHEET, index=False)
            return True
    except Exception as e:
        st.error(f"Failed to delete recipe: {e}")
        return False

# Helper Function: Generate Automatic Recipe ID
def get_next_recipe_id():
    df_recipes = load_saved_recipes()
    year = datetime.now().year
    if df_recipes.empty:
        return f"RND-REC-{year}-001"
    return f"RND-REC-{year}-{len(df_recipes) + 1:03d}"

# Initialize Session State
if "ink_library" not in st.session_state:
    st.session_state.ink_library = load_ink_library()

if "formulation_df" not in st.session_state:
    st.session_state.formulation_df = pd.DataFrame([
        {"Delete": False, "Role": "Base Transparent", "Product Name": "Silicone Base Clear", "Code": "SIL-BASE-90", "Percentage (%)": 70.0, "Unit Price ($/kg)": 14.50, "Mixing Notes": "High elasticity base"},
        {"Delete": False, "Role": "Base Opaque / White", "Product Name": "Silicone White Undercoat", "Code": "SIL-WHT-10", "Percentage (%)": 20.0, "Unit Price ($/kg)": 12.00, "Mixing Notes": "Provides opacity & bleed barrier"},
        {"Delete": False, "Role": "Pigment Colorant", "Product Name": "High Fastness Black Pigment", "Code": "PIG-BLK-05", "Percentage (%)": 5.0, "Unit Price ($/kg)": 18.00, "Mixing Notes": "Color shade matching"},
        {"Delete": False, "Role": "Catalyst / Hardener", "Product Name": "Silicone Platinum Catalyst", "Code": "CAT-SIL-02", "Percentage (%)": 2.0, "Unit Price ($/kg)": 45.00, "Mixing Notes": "Mix thoroughly before printing"},
        {"Delete": False, "Role": "Crosslinker / Fixer", "Product Name": "Anti-Fading Crosslinker", "Code": "XL-MOD-01", "Percentage (%)": 3.0, "Unit Price ($/kg)": 28.00, "Mixing Notes": "Enhances wash fastness (20+ cycles)"}
    ])

# Helper Function: Load Master Logs
def load_data():
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame(), pd.DataFrame()
    try:
        xls = pd.ExcelFile(EXCEL_FILE)
        trials_df = pd.read_excel(xls, "R&D Master Trial Log", skiprows=2) if "R&D Master Trial Log" in xls.sheet_names else pd.DataFrame()
        matrix_df = pd.read_excel(xls, "Fabric Compatibility Matrix", skiprows=2) if "Fabric Compatibility Matrix" in xls.sheet_names else pd.DataFrame()
        return trials_df, matrix_df
    except Exception as e:
        st.error(f"Error loading workbook: {e}")
        return pd.DataFrame(), pd.DataFrame()

# Helper Function: Save Trial to Excel
def save_trial_to_excel(trial_data):
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "R&D Master Trial Log"
            ws.append([])
            ws.append([])
            ws.append(["Trial ID", "Date", "Style / Reference", "Technique", "Fabric Type", "Objective", "Recipe Variation", "Wash Result", "Crocking Result", "Status"])
        else:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb["R&D Master Trial Log"] if "R&D Master Trial Log" in wb.sheetnames else wb.create_sheet("R&D Master Trial Log")
            if ws.max_row == 1:
                ws.append([])
                ws.append([])
                ws.append(["Trial ID", "Date", "Style / Reference", "Technique", "Fabric Type", "Objective", "Recipe Variation", "Wash Result", "Crocking Result", "Status"])
        
        ws.append(list(trial_data.values()))
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Failed to save to Excel file: {e}")
        return False

# PDF Generator Function
def generate_pdf_report(data):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=25, 
        leftMargin=25, 
        topMargin=25, 
        bottomMargin=35
    )
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('HeaderTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=14, leading=16, textColor=colors.HexColor('#1E293B'), alignment=1)
    subtitle_style = ParagraphStyle('HeaderSub', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.HexColor('#2563EB'), alignment=1)
    sec_style = ParagraphStyle('SecTitle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=colors.HexColor('#2563EB'), spaceBefore=6, spaceAfter=3)
    p_style = ParagraphStyle('BodyCustom', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10)
    
    elements = []
    elements.append(Paragraph("JK GARMENT SCREEN PRINTING R&D", title_style))
    elements.append(Paragraph("TECHNICAL RECIPE & PRODUCTION REQUIREMENT REPORT", subtitle_style))
    elements.append(Spacer(1, 4))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#2563EB'), spaceAfter=6))
    
    # Header Info
    h_data = [
        [Paragraph("<b>Recipe ID:</b>", p_style), Paragraph(str(data.get('recipe_id', '')), p_style), Paragraph("<b>Date:</b>", p_style), Paragraph(str(data.get('date', '')), p_style)],
        [Paragraph("<b>Style Name/No:</b>", p_style), Paragraph(str(data.get('style_name', '')), p_style), Paragraph("<b>Print Tech:</b>", p_style), Paragraph(str(data.get('print_tech', '')), p_style)],
        [Paragraph("<b>Created by:</b>", p_style), Paragraph(str(data.get('created_by', '')), p_style), Paragraph("<b>Revision:</b>", p_style), Paragraph(str(data.get('revision', '')), p_style)],
        [Paragraph("<b>Fabric Comp:</b>", p_style), Paragraph(str(data.get('fabric_comp', '')), p_style), Paragraph("<b>Fabric Color:</b>", p_style), Paragraph(str(data.get('fabric_color', '')), p_style)],
        [Paragraph("<b>Fabric Const:</b>", p_style), Paragraph(str(data.get('fabric_const', '')), p_style), Paragraph("<b>GSM:</b>", p_style), Paragraph(str(data.get('gsm', '')), p_style)],
        [Paragraph("<b>Dye Migration Risk:</b>", p_style), Paragraph(str(data.get('dye_risk', '')), p_style), Paragraph("<b>Undercoat Req:</b>", p_style), Paragraph(str(data.get('undercoat', '')), p_style)]
    ]
    t_h = Table(h_data, colWidths=[90, 180, 90, 180])
    t_h.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')), 
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')), 
        ('PADDING', (0,0), (-1,-1), 3)
    ]))
    elements.append(t_h)
    
    # Bulk Production & Cost Planning
    elements.append(Paragraph("1. BULK PRODUCTION & MATERIAL PLANNING (RM REQUIREMENT)", sec_style))
    p_data = [
        [Paragraph("<b>Target Sample Pcs:</b>", p_style), Paragraph(f"{data.get('target_pcs', 0)} pcs", p_style), Paragraph("<b>Order Quantity:</b>", p_style), Paragraph(f"{data.get('order_qty', 0):,} pcs", p_style)],
        [Paragraph("<b>Batch Size:</b>", p_style), Paragraph(f"{data.get('batch_size', 0)} g", p_style), Paragraph("<b>Wastage Allowance:</b>", p_style), Paragraph(f"{data.get('wastage_pct', 0)}%", p_style)],
        [Paragraph("<b>Ink Used / Pc:</b>", p_style), Paragraph(f"{float(data.get('per_pc_used', 0)):.2f} g/pc", p_style), Paragraph("<b>Total Bulk RM Req:</b>", p_style), Paragraph(f"<b>{float(data.get('total_bulk_rm_kg', 0)):.2f} kg</b>", p_style)],
        [Paragraph("<b>Estimated Recipe Cost:</b>", p_style), Paragraph(f"${float(data.get('cost_per_kg', 0)):.2f} / kg", p_style), Paragraph("<b>Cost Per Garment:</b>", p_style), Paragraph(f"<b>${float(data.get('cost_per_pc', 0)):.4f} / pc</b>", p_style)]
    ]
    t_p = Table(p_data, colWidths=[110, 160, 110, 160])
    t_p.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')), 
        ('PADDING', (0,0), (-1,-1), 3)
    ]))
    elements.append(t_p)

    # Formulation Table
    elements.append(Paragraph("2. INK FORMULATION & BOM COST BREAKDOWN", sec_style))
    ink_headers = ["Role", "Product Name", "Code", "Ratio", "Sample (g)", "Per Pc (g)", "Bulk Req (kg)", "Price/kg", "Total ($)"]
    i_table = [[Paragraph(f"<b>{h}</b>", ParagraphStyle('TH', parent=p_style, textColor=colors.white)) for h in ink_headers]]
    
    formulation_list = json.loads(data['formulation']) if isinstance(data.get('formulation'), str) else data.get('formulation', [])
    for row in formulation_list:
        i_table.append([
            Paragraph(str(row.get('Role', '')), p_style),
            Paragraph(str(row.get('Product Name', '')), p_style),
            Paragraph(str(row.get('Code', '')), p_style),
            Paragraph(f"{float(row.get('Percentage (%)', 0)):.1f}%", p_style),
            Paragraph(f"{float(row.get('Use Quantity (g)', 0)):.1f}g", p_style),
            Paragraph(f"{float(row.get('Per 1 Used (g)', 0)):.2f}g", p_style),
            Paragraph(f"{float(row.get('Bulk Req (kg)', 0)):.2f}kg", p_style),
            Paragraph(f"${float(row.get('Unit Price ($/kg)', 0)):.2f}", p_style),
            Paragraph(f"${float(row.get('Line Cost ($)', 0)):.2f}", p_style),
        ])
    t_i = Table(i_table, colWidths=[70, 105, 50, 40, 50, 50, 55, 50, 50])
    t_i.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')), 
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')), 
        ('PADDING', (0,0), (-1,-1), 3)
    ]))
    elements.append(t_i)

    # Technical Machine Parameters Section
    elements.append(Paragraph("3. TECHNICAL MACHINE PARAMETERS", sec_style))
    m_data = [
        [Paragraph("<b>Mesh Count:</b>", p_style), Paragraph(str(data.get('mesh', '')), p_style), Paragraph("<b>Flash Cure Temp/Time:</b>", p_style), Paragraph(str(data.get('flash_cure', '')), p_style)],
        [Paragraph("<b>Squeegee Durometer:</b>", p_style), Paragraph(str(data.get('squeegee_duro', '')), p_style), Paragraph("<b>Main Curing Temp:</b>", p_style), Paragraph(str(data.get('main_cure', '')), p_style)],
        [Paragraph("<b>Squeegee Angle/Speed:</b>", p_style), Paragraph(str(data.get('squeegee_angle', '')), p_style), Paragraph("<b>Drying Belt Speed:</b>", p_style), Paragraph(str(data.get('belt_speed', '')), p_style)],
        [Paragraph("<b>Off-Contact Distance:</b>", p_style), Paragraph(str(data.get('off_contact', '')), p_style), Paragraph("<b>Number of Passes:</b>", p_style), Paragraph(str(data.get('passes', '')), p_style)],
    ]
    t_m = Table(m_data, colWidths=[110, 160, 110, 160])
    t_m.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')), 
        ('PADDING', (0,0), (-1,-1), 3)
    ]))
    elements.append(t_m)
    
    # Sign-off Table
    elements.append(Paragraph("4. TECHNICAL APPROVAL", sec_style))
    s_data = [
        [Paragraph("<b>R&D Exec:</b>", p_style), Paragraph(str(data.get('sig_rd', '')), p_style), Paragraph("<b>Quality Dept:</b>", p_style), Paragraph(str(data.get('sig_qa', '')), p_style)],
        [Paragraph("<b>Sample Dev:</b>", p_style), Paragraph(str(data.get('sig_sample', '')), p_style), Paragraph("<b>Production Mgr:</b>", p_style), Paragraph(str(data.get('sig_prod', '')), p_style)],
    ]
    t_s = Table(s_data, colWidths=[90, 185, 90, 185])
    t_s.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')), 
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')), 
        ('PADDING', (0,0), (-1,-1), 3)
    ]))
    elements.append(t_s)

    # Footer Callback Function for Created By & Developed By Labels
    created_by_text = str(data.get('created_by', '')).strip().lower()
    left_footer = f"created by {created_by_text}" if created_by_text else "created by unknown"
    right_footer = "Developed by - Lab Executive - Lakshan Vimukthi"

    def draw_footer(canvas, pdf_doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor('#64748B'))
        # Bottom-Left: Small letters creator name
        canvas.drawString(25, 15, left_footer)
        # Bottom-Right: Small words Developed by - Lab Executive - Lakshan Vimukthi
        canvas.drawRightString(pdf_doc.pagesize[0] - 25, 15, right_footer)
        canvas.restoreState()

    doc.build(elements, onFirstPage=draw_footer, onLaterPages=draw_footer)
    buffer.seek(0)
    return buffer.getvalue()

st.title("🎨 Screen Printing R&D & Costing Portal")

menu = st.sidebar.radio(
    "Select Action", 
    [
        "🎨 Technical Recipe Builder & Report Generator", 
        "📖 View & Search Saved Recipes",
        "📚 Chemical & Ink Library Manager",
        "🧪 Log R&D Trial Result", 
        "📊 View Master Trial Log", 
        "🧩 Fabric Compatibility Matrix"
    ]
)

if menu == "🎨 Technical Recipe Builder & Report Generator":
    st.header("1. Job Header & Substrate Specifications")
    auto_recipe_id = get_next_recipe_id()

    col1, col2 = st.columns(2)
    with col1:
        recipe_id = st.text_input("Recipe ID (Auto-Generated)", value=auto_recipe_id, disabled=True)
        style_name = st.text_input("Style Name / No", value="", placeholder="e.g. ST-2026-GYMSHARK-01")
        created_by = st.text_input("Created by", value="", placeholder="e.g. John Doe / Lab Exec")
        fabric_comp = st.text_input("Fabric Composition", value="", placeholder="e.g. 95% Cotton / 5% Elastane")
        fabric_const = st.text_input("Fabric Construction", value="", placeholder="e.g. Single Jersey (Knitted)")
        dye_migration = st.selectbox("Dye Migration Risk", ["Low", "Medium", "High", "Critical"], index=1)
    
    with col2:
        rec_date = st.date_input("Date", value=datetime.now().date())
        print_tech = st.text_input("Print Technique", value="", placeholder="e.g. Silicone Rubber, High Density")
        revision = st.text_input("Revision No", value="v2.1", placeholder="e.g. v1.0, v2.1")
        fabric_color = st.text_input("Fabric Color / CW", value="", placeholder="e.g. Charcoal Dark Grey")
        gsm = st.text_input("GSM", value="", placeholder="e.g. 180 GSM, 220 GSM")
        undercoat = st.selectbox("Undercoat Required", ["Yes", "No"], index=1)

    st.divider()
    st.header("2. Production Metrics & Material Requirement Planning (RM)")
    
    prod_col1, prod_col2, prod_col3, prod_col4 = st.columns(4)
    with prod_col1:
        batch_size = st.number_input("Sample Batch Size (Grams)", value=1000.0, step=50.0, min_value=1.0)
    with prod_col2:
        target_pcs = st.number_input("Target Printed Pcs (From Batch)", value=50, step=5, min_value=1)
    with prod_col3:
        order_qty = st.number_input("Bulk Order Qty (Garments)", value=5000, step=500, min_value=1)
    with prod_col4:
        wastage_pct = st.number_input("Wastage Allowance (%)", value=10.0, step=1.0, min_value=0.0)

    # Derived production values
    per_pc_used = batch_size / target_pcs  # Grams per garment printed
    wastage_factor = 1.0 + (wastage_pct / 100.0)
    total_bulk_rm_kg = (order_qty * per_pc_used * wastage_factor) / 1000.0

    st.info(f"💡 **Calculated Ink/Garment:** `{per_pc_used:.2f} g/pc` | **Total Bulk Ink Requirement (incl. {wastage_pct}% Wastage):** `{total_bulk_rm_kg:.2f} kg`")

    st.divider()
    st.header("3. Chemical Recipe Formulation & Costing")

    # Sync price and code mappings from master library
    ink_lib = st.session_state.ink_library
    code_map = dict(zip(ink_lib["Product Name"], ink_lib["Supplier Code"]))
    price_map = dict(zip(ink_lib["Product Name"], ink_lib["Unit Price ($/kg)"]))
    available_products = sorted(ink_lib["Product Name"].dropna().tolist()) or [""]

    form_df = st.session_state.formulation_df.copy()

    # Synchronize codes & prices
    for idx, row in form_df.iterrows():
        pname = row.get("Product Name")
        if pname in code_map:
            form_df.at[idx, "Code"] = code_map[pname]
        if pname in price_map and (pd.isna(row.get("Unit Price ($/kg)")) or row.get("Unit Price ($/kg)") == 0):
            form_df.at[idx, "Unit Price ($/kg)"] = price_map[pname]

    column_order = ["Delete", "Role", "Product Name", "Code", "Percentage (%)", "Unit Price ($/kg)", "Mixing Notes"]
    display_df = form_df.reindex(columns=column_order)

    edited_df = st.data_editor(
        display_df,
        num_rows="dynamic",
        column_config={
            "Delete": st.column_config.CheckboxColumn("Delete?", default=False),
            "Role": st.column_config.SelectboxColumn("Component Role", options=["Base Transparent", "Base Opaque / White", "Pigment Colorant", "Catalyst / Hardener", "Crosslinker / Fixer", "Additive / Retarder", "Thinner / Reducer", "Thickeners", "Other"], required=True),
            "Product Name": st.column_config.SelectboxColumn("Product Name", options=available_products, required=True),
            "Code": st.column_config.TextColumn("Code", disabled=True),
            "Percentage (%)": st.column_config.NumberColumn("Ratio (%)", min_value=0.0, max_value=100.0, step=0.5, format="%.1f%%"),
            "Unit Price ($/kg)": st.column_config.NumberColumn("Unit Price ($/kg)", min_value=0.0, step=0.50, format="$%.2f"),
            "Mixing Notes": st.column_config.TextColumn("Mixing Notes")
        },
        use_container_width=True,
        key="formulation_editor"
    )

    st.session_state.formulation_df = edited_df

    if st.button("🗑️ Delete Selected Rows"):
        if "Delete" in edited_df.columns:
            st.session_state.formulation_df = edited_df[edited_df["Delete"] == False].reset_index(drop=True)
            st.rerun()

    # Dynamic Calculations
    calc_df = st.session_state.formulation_df.copy()
    calc_df["Percentage (%)"] = pd.to_numeric(calc_df["Percentage (%)"], errors="coerce").fillna(0.0)
    calc_df["Unit Price ($/kg)"] = pd.to_numeric(calc_df["Unit Price ($/kg)"], errors="coerce").fillna(0.0)

    calc_df["Use Quantity (g)"] = (calc_df["Percentage (%)"] / 100.0) * batch_size
    calc_df["Per 1 Used (g)"] = (calc_df["Percentage (%)"] / 100.0) * per_pc_used
    calc_df["Bulk Req (kg)"] = (calc_df["Percentage (%)"] / 100.0) * total_bulk_rm_kg
    calc_df["Line Cost ($)"] = calc_df["Bulk Req (kg)"] * calc_df["Unit Price ($/kg)"]

    total_pct = calc_df["Percentage (%)"].sum()
    total_batch_weight = calc_df["Use Quantity (g)"].sum()
    total_cost_usd = calc_df["Line Cost ($)"].sum()
    cost_per_kg = (total_cost_usd / total_bulk_rm_kg) if total_bulk_rm_kg > 0 else 0.0
    cost_per_pc = total_cost_usd / order_qty if order_qty > 0 else 0.0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Formulation Ratio", f"{total_pct:.1f}%")
    m2.metric("Total Batch Weight", f"{total_batch_weight:.1f} g")
    m3.metric("Cost per kg of Ink", f"${cost_per_kg:.2f} / kg")
    m4.metric("Print Cost per Garment", f"${cost_per_pc:.4f} / pc")

    st.subheader("Calculated Requirement & Costing Preview")
    st.dataframe(
        calc_df[["Role", "Product Name", "Code", "Percentage (%)", "Use Quantity (g)", "Per 1 Used (g)", "Bulk Req (kg)", "Unit Price ($/kg)", "Line Cost ($)"]],
        column_config={
            "Percentage (%)": st.column_config.NumberColumn("Ratio (%)", format="%.1f%%"),
            "Use Quantity (g)": st.column_config.NumberColumn("Batch Qty (g)", format="%.1f g"),
            "Per 1 Used (g)": st.column_config.NumberColumn("Usage/Pc (g)", format="%.2f g"),
            "Bulk Req (kg)": st.column_config.NumberColumn("Bulk RM Req (kg)", format="%.2f kg"),
            "Unit Price ($/kg)": st.column_config.NumberColumn("Price ($/kg)", format="$%.2f"),
            "Line Cost ($)": st.column_config.NumberColumn("Total Cost ($)", format="$%.2f")
        },
        use_container_width=True
    )

    st.divider()
    st.header("4. Technical Machine Parameters & Sign-off")
    m_col1, m_col2 = st.columns(2)
    with m_col1:
        mesh = st.text_input("Mesh Count (T/Inch)", value="", placeholder="e.g. 120T (305 Mesh)")
        squeegee_duro = st.text_input("Squeegee Durometer", value="", placeholder="e.g. 70/90/70 Triple")
        squeegee_angle = st.text_input("Squeegee Angle / Speed", value="", placeholder="e.g. 75° / Medium Speed")
        off_contact = st.text_input("Off-Contact Distance", value="", placeholder="e.g. 2.5 mm")
    with m_col2:
        flash_cure = st.text_input("Flash Cure Temp / Time", value="", placeholder="e.g. 110°C / 5 Seconds")
        main_cure = st.text_input("Main Curing Temp", value="", placeholder="e.g. 100°C")
        belt_speed = st.text_input("Drying Belt Speed / Time", value="", placeholder="e.g. 90sec")
        passes = st.text_input("Number of Passes / Strokes", value="", placeholder="e.g. 2 Print - 1 Flash - 2 Print")

    s_col1, s_col2 = st.columns(2)
    with s_col1:
        sig_rd = st.selectbox("R&D Exec Approval", ["Pending", "Approved", "Rejected"], index=1)
        sig_sample = st.selectbox("Sample Dev Head Approval", ["Pending", "Approved", "Rejected"], index=1)
    with s_col2:
        sig_qa = st.selectbox("Quality Dept Approval", ["Pending", "Approved", "Rejected"], index=1)
        sig_prod = st.selectbox("Production Manager Approval", ["Pending", "Approved", "Rejected"], index=1)

    recipe_summary = {
        'recipe_id': recipe_id, 'date': str(rec_date), 'style_name': style_name,
        'print_tech': print_tech, 'created_by': created_by, 'revision': revision,
        'fabric_comp': fabric_comp, 'fabric_color': fabric_color, 'fabric_const': fabric_const,
        'gsm': gsm, 'dye_risk': dye_migration, 'undercoat': undercoat,
        'batch_size': batch_size, 'target_pcs': target_pcs, 'order_qty': order_qty,
        'wastage_pct': wastage_pct, 'per_pc_used': per_pc_used, 'total_bulk_rm_kg': total_bulk_rm_kg,
        'cost_per_kg': cost_per_kg, 'cost_per_pc': cost_per_pc,
        'formulation': calc_df.to_dict(orient="records"),
        'mesh': mesh, 'squeegee_duro': squeegee_duro, 'squeegee_angle': squeegee_angle,
        'off_contact': off_contact, 'flash_cure': flash_cure, 'main_cure': main_cure,
        'belt_speed': belt_speed, 'passes': passes, 'sig_rd': sig_rd, 'sig_qa': sig_qa,
        'sig_sample': sig_sample, 'sig_prod': sig_prod
    }

    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        if st.button("💾 Save Recipe & Material Plan to Excel"):
            if save_recipe_to_excel(recipe_summary):
                st.success(f"✅ Recipe '{recipe_id}' saved successfully!")
                st.rerun()

    with btn_col2:
        if st.button("📄 Generate PDF Specification & Cost Report"):
            pdf_bytes = generate_pdf_report(recipe_summary)
            st.success("✅ PDF Specification Report generated!")
            st.download_button(
                label="💾 Download Technical PDF",
                data=pdf_bytes,
                file_name=f"Recipe_Report_{recipe_id}.pdf",
                mime="application/pdf"
            )

elif menu == "📖 View & Search Saved Recipes":
    st.header("📖 Saved Technical Recipes & Cost Library")
    df_recipes = load_saved_recipes()

    if df_recipes.empty:
        st.info("No saved recipes found. Create and save a new recipe from the Recipe Builder!")
    else:
        st.dataframe(df_recipes[['recipe_id', 'date', 'style_name', 'print_tech', 'order_qty', 'total_bulk_rm_kg', 'cost_per_pc']], use_container_width=True)
        st.divider()
        selected_id = st.selectbox("Select Recipe ID to Inspect", options=df_recipes['recipe_id'].tolist())

        if selected_id:
            recipe_row = df_recipes[df_recipes['recipe_id'] == selected_id].iloc[0].to_dict()
            st.markdown(f"### Details for `{recipe_row['recipe_id']}`")
            col1, col2, col3 = st.columns(3)
            col1.write(f"**Style Name:** {recipe_row['style_name']}")
            col1.write(f"**Order Qty:** {recipe_row.get('order_qty', 0):,} pcs")
            col2.write(f"**Print Tech:** {recipe_row['print_tech']}")
            col2.write(f"**Bulk Requirement:** {recipe_row.get('total_bulk_rm_kg', 0):.2f} kg")
            col3.write(f"**Cost/Pc:** ${recipe_row.get('cost_per_pc', 0):.4f}")

            form_data = json.loads(recipe_row['formulation']) if isinstance(recipe_row['formulation'], str) else recipe_row['formulation']
            st.dataframe(pd.DataFrame(form_data), use_container_width=True)

            action_col1, action_col2 = st.columns(2)
            with action_col1:
                pdf_bytes = generate_pdf_report(recipe_row)
                st.download_button(label=f"📄 Download PDF", data=pdf_bytes, file_name=f"Report_{selected_id}.pdf", mime="application/pdf")

            with action_col2:
                if st.button(f"Permanently Delete {selected_id}", type="primary"):
                    if delete_recipe_from_excel(selected_id):
                        st.success(f"Deleted {selected_id} successfully!")
                        st.rerun()

elif menu == "📚 Chemical & Ink Library Manager":
    st.header("📚 Chemical & Advanced Ink Library")
    
    with st.expander("➕ Add New Chemical Product to Library"):
        with st.form("add_chem_form"):
            new_pname = st.text_input("Product Name", placeholder="e.g. Silicone Base Clear")
            new_code = st.text_input("Supplier / Product Code", placeholder="e.g. SIL-BASE-90")
            new_role = st.selectbox("Component Role", ["Base Transparent", "Base Opaque / White", "Pigment Colorant", "Catalyst / Hardener", "Crosslinker / Fixer", "Additive / Retarder", "Thinner / Reducer"])
            new_price = st.number_input("Unit Price ($/kg)", min_value=0.0, step=0.50, value=12.00)
            new_notes = st.text_area("Notes & Properties", placeholder="e.g. High elasticity silicone base")
            
            if st.form_submit_button("Add to Library"):
                if new_pname and new_code:
                    new_item = pd.DataFrame([{"Product Name": new_pname, "Supplier Code": new_code, "Role": new_role, "Unit Price ($/kg)": new_price, "Notes": new_notes}])
                    st.session_state.ink_library = pd.concat([st.session_state.ink_library, new_item], ignore_index=True)
                    if save_ink_library_to_excel(st.session_state.ink_library):
                        st.success(f"Added '{new_pname}' to chemical library!")
                        st.rerun()
                else:
                    st.warning("Please provide Product Name and Code.")

    st.dataframe(st.session_state.ink_library, use_container_width=True)

elif menu == "🧪 Log R&D Trial Result":
    st.header("Log Experimental & Durability Trial Result")
    with st.form("log_trial_form"):
        col1, col2 = st.columns(2)
        with col1:
            trial_id = st.text_input("Trial ID", value="TR-2026-006")
            style_ref = st.text_input("Style / Reference", value="ST-2026-GYM-01")
            fabric_type = st.text_input("Fabric Type", value="95% Ctn / 5% Elastane")
            recipe_var = st.text_area("Recipe / Parameter Variation", value="Added 3% Crosslinker XL-MOD-01")
            crocking_res = st.text_input("Crocking Result", value="Grade 4.5")
        with col2:
            trial_date = st.date_input("Date", value=datetime.now().date())
            technique = st.selectbox("Technique", ["Silicone", "Rubber Print", "High Density", "Flock Print", "Glitter Print"])
            objective = st.text_area("Trial Objective", value="Improve wash fastness past 20 cycles")
            wash_res = st.text_input("Wash Result (20 Cycles)", value="Grade 4.5 (Pass)")
            status = st.selectbox("Overall Status", ["APPROVED", "REJECTED", "PENDING"])
        
        if st.form_submit_button("Save Trial Log"):
            trial_data = {"Trial ID": trial_id, "Date": str(trial_date), "Style / Reference": style_ref, "Technique": technique, "Fabric Type": fabric_type, "Objective": objective, "Recipe Variation": recipe_var, "Wash Result": wash_res, "Crocking Result": crocking_res, "Status": status}
            if save_trial_to_excel(trial_data):
                st.success(f"Trial {trial_id} saved successfully!")

elif menu == "📊 View Master Trial Log":
    st.header("R&D Master Experimental & Durability Trial Log")
    fresh_trials, _ = load_data()
    st.dataframe(fresh_trials, use_container_width=True) if not fresh_trials.empty else st.info("No trial records found.")

elif menu == "🧩 Fabric Compatibility Matrix":
    st.header("Print Technique vs. Fabric Substrate Compatibility Matrix")
    _, fresh_matrix = load_data()
    st.dataframe(fresh_matrix, use_container_width=True) if not fresh_matrix.empty else st.info("No compatibility matrix sheet found.")